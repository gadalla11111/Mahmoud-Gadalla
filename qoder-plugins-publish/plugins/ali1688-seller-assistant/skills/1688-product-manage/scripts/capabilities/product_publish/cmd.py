#!/usr/bin/env python3
"""商品发布命令 — CLI 入口"""

from __future__ import annotations

COMMAND_NAME = "product_publish"
COMMAND_DESC = "图片发布商品（单图/批量，支持本地图片或图片链接）"

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))

import argparse
import io
import json
import platform
import subprocess
import tempfile
import time
import urllib.request
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Font, Alignment, Protection

from PIL import Image as PILImage

from _output import make_output, print_output, print_error
from capabilities.product_publish.service import (
    upload_image_to_bank,
    ai_publish_by_image,
    ai_publish_save,
    build_publish_url,
)

# 状态文件：项目根目录，用于在多次 CLI 调用间保持发品上下文
STATE_FILE = os.path.join(os.path.dirname(__file__), "..", "..", ".publish_state.json")
BATCH_STATE_FILE = os.path.join(os.path.dirname(__file__), "..", "..", ".batch_publish_state.json")


def _save_state(state: dict) -> None:
    state["saved_at"] = time.time()
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False)


def _load_state() -> dict | None:
    if not os.path.exists(STATE_FILE):
        return None
    try:
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def _save_batch_state(items: list) -> None:
    state = {"items": items, "saved_at": time.time()}
    with open(BATCH_STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False)


def _load_batch_state() -> dict | None:
    if not os.path.exists(BATCH_STATE_FILE):
        return None
    try:
        with open(BATCH_STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


# 图片插入配置
_THUMB_SIZE = (120, 120)   # 缩略图最大尺寸（像素）
_THUMB_ROW_HEIGHT = 94     # 行高（points）≈ 125px，与列宽接近方形
_IMG_COL_WIDTH = 17        # 本地图片列宽（字符）≈ 129px


def _prepare_image_for_excel(pic_path: str, pic_url: str, row_idx: int) -> str:
    """生成缩略图临时文件并返回路径；失败时返回空字符串。"""
    try:
        if os.path.isfile(pic_path):
            img = PILImage.open(pic_path)
        elif pic_url:
            data = urllib.request.urlopen(pic_url, timeout=15).read()
            img = PILImage.open(io.BytesIO(data))
        else:
            return ""
        img.thumbnail(_THUMB_SIZE, PILImage.LANCZOS)
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
        tmp_path = os.path.join(tempfile.gettempdir(), f"_excel_thumb_{row_idx}.jpg")
        img.save(tmp_path, format="JPEG", quality=85)
        return tmp_path
    except Exception:
        return ""


def _format_seller_point(seller_point: dict) -> str:
    """将 sameItemSellerPoint dict 格式化为可读字符串。"""
    if not seller_point or not isinstance(seller_point, dict):
        return "无"
    return ", ".join(f"{k}: {v}" for k, v in seller_point.items())


# HTML 输出目录：scripts/ 目录下
HTML_OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..")


def _generate_select_same_html(candidates: list) -> str:
    """生成同款勾选交互HTML文件，返回文件绝对路径。

    参数 candidates 格式:
    [
        {
            "index": 1,                    # 图片序号
            "main_image_url": "https://...",  # 商家原始主图URL
            "same_items": [                # 同款候选列表
                {
                    "seq": 1,              # 同款序号
                    "image_url": "https://...",  # 同款主图URL
                    "title": "xxx",        # 同款标题
                    "attributes": "颜色:红色 尺码:XL"  # 属性键值对
                },
                ...
            ]
        },
        ...
    ]
    """
    import html as html_mod

    # 构建 JSON 数据嵌入 HTML
    candidates_json = json.dumps(candidates, ensure_ascii=False)

    html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<title>批量发品 - 同款勾选</title>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: -apple-system, "PingFang SC", "Microsoft YaHei", sans-serif; padding: 24px; background: #f5f7fa; color: #333; }}
.header {{ max-width: 1200px; margin: 0 auto 12px; }}
.header h1 {{ font-size: 1.4rem; margin-bottom: 8px; }}
.header p {{ color: #666; font-size: 0.9rem; }}
.tip {{ max-width: 1200px; margin: 0 auto 16px; padding: 10px 14px; background: #fff7e6; border: 1px solid #ffd591; border-radius: 6px; font-size: 0.85rem; color: #874d00; line-height: 1.7; }}
.tip b {{ color: #d46b08; }}
.toolbar {{ max-width: 1200px; margin: 16px auto; display: flex; gap: 12px; flex-wrap: wrap; }}
.toolbar button {{ padding: 8px 16px; border: 1px solid #ddd; border-radius: 6px; background: #fff; cursor: pointer; font-size: 0.85rem; transition: all 0.2s; }}
.toolbar button:hover {{ background: #FF6A00; color: #fff; border-color: #FF6A00; }}
.toolbar button.primary {{ background: #FF6A00; color: #fff; border-color: #FF6A00; }}
.toolbar button.primary:hover {{ background: #e55d00; }}
.group-card {{ max-width: 1200px; margin: 0 auto 16px; background: #fff; border-radius: 8px; box-shadow: 0 2px 12px rgba(0,0,0,0.06); overflow: hidden; }}
.main-image-row {{ display: flex; align-items: center; gap: 12px; padding: 12px 16px; background: #f8f9fa; border-bottom: 1px solid #eee; }}
.main-image-row img {{ width: 120px; height: 120px; object-fit: cover; border-radius: 6px; border: 1px solid #eee; }}
.main-image-row .label {{ font-size: 1rem; font-weight: 600; color: #333; white-space: nowrap; }}
.same-items-table {{ width: 100%; border-collapse: collapse; }}
.same-items-table th {{ padding: 10px 10px; text-align: left; font-size: 0.85rem; color: #555; border-bottom: 2px solid #eee; background: #fafbfc; }}
.same-items-table td {{ padding: 10px; border-bottom: 1px solid #f0f0f0; vertical-align: middle; font-size: 0.85rem; }}
.same-items-table tr:last-child td {{ border-bottom: none; }}
.same-items-table .img-cell img {{ width: 80px; height: 80px; object-fit: cover; border-radius: 4px; border: 1px solid #eee; }}
.same-items-table .check-cell {{ text-align: center; }}
.same-items-table .check-cell input[type="checkbox"] {{ width: 18px; height: 18px; cursor: pointer; accent-color: #FF6A00; }}
.toast {{ position: fixed; top: 20px; left: 50%; transform: translateX(-50%); background: #333; color: #fff; padding: 10px 24px; border-radius: 6px; font-size: 0.9rem; opacity: 0; transition: opacity 0.3s; pointer-events: none; z-index: 999; }}
.toast.show {{ opacity: 1; }}
.attr-empty-val {{ color: #ccc; text-align: center; }}
.table-wrap {{ overflow-x: auto; }}
.same-items-table th.grp-p1 {{ background: #ffe7ba; color: #874d00; font-weight: 600; text-align: center; }}
.same-items-table th.grp-p2 {{ background: #d6e4ff; color: #1d39c4; font-weight: 600; text-align: center; }}
.same-items-table th.grp-p3 {{ background: #efdbff; color: #531dab; font-weight: 600; text-align: center; }}
.same-items-table th.col-p1, .same-items-table td.cell-p1 {{ background: #fffbe6; }}
.same-items-table th.col-p2, .same-items-table td.cell-p2 {{ background: #f5f9ff; }}
.same-items-table th.col-p3, .same-items-table td.cell-p3 {{ background: #fcf4ff; }}
.same-items-table th.col-key {{ font-size: 0.78rem; min-width: 80px; max-width: 200px; word-break: break-word; padding: 6px 8px; color: #555; }}
.same-items-table td.title-cell {{ min-width: 160px; max-width: 240px; }}
.diff-hint {{ font-size: 0.8rem; padding: 3px 10px; border-radius: 4px; margin-left: 12px; background: #fafafa; border: 1px solid #e8e8e8; }}
.diff-hint .sum-p1 {{ color: #d46b08; }}
.diff-hint .sum-p2 {{ color: #1d39c4; }}
.diff-hint .sum-p3 {{ color: #531dab; }}
.diff-hint b {{ font-weight: 700; }}
.no-diff {{ font-size: 0.8rem; color: #52c41a; margin-left: 12px; }}
</style>
</head>
<body>
<div class="header">
  <h1>🛒 批量发品 - 同款勾选</h1>
</div>

<div class="tip">
  💡 <b>使用说明</b>：每张图片<b>必须选择一个</b>同款（不可取消）；点击其他同款可切换选择。确认后点击下方按钮复制结果。
</div>

<div class="toolbar">
  <button class="primary" onclick="copySelections()">📋 复制勾选结果</button>
</div>

<div id="cardsContainer"></div>

<div class="toolbar">
  <button class="primary" onclick="copySelections()">📋 复制勾选结果</button>
</div>

<div class="toast" id="toast"></div>

<script>
const CANDIDATES = {candidates_json};

function escapeHtml(s) {{
  return String(s == null ? '' : s)
    .replace(/&/g, '&amp;')
    .replace(/</g, '&lt;')
    .replace(/>/g, '&gt;')
    .replace(/"/g, '&quot;');
}}

function parseAttributes(attrStr) {{
  if (!attrStr) return [];
  const pairs = String(attrStr).split(/[,，]\s*/).map(s => s.trim()).filter(Boolean);
  return pairs.map(pair => {{
    const m = pair.match(/^(.+?)[:：]\s*(.*)$/);
    if (!m) return {{ key: pair, value: '' }};
    return {{ key: m[1].trim(), value: m[2].trim() }};
  }});
}}

function classifyAttributes(items) {{
  const maps = items.map(item => {{
    const m = {{}};
    parseAttributes(item.attributes).forEach(({{key, value}}) => {{
      m[key] = value;
    }});
    return m;
  }});
  const keyCounts = {{}};
  const keyValues = {{}};
  maps.forEach(m => {{
    Object.keys(m).forEach(k => {{
      keyCounts[k] = (keyCounts[k] || 0) + 1;
      if (!keyValues[k]) keyValues[k] = new Set();
      keyValues[k].add(m[k]);
    }});
  }});
  const total = items.length;
  const p1 = [], p2 = [], p3 = [];
  Object.keys(keyCounts).forEach(k => {{
    const cnt = keyCounts[k];
    if (cnt === total) {{
      // 全部同款都有该键：值不全一致 -> p1，值全一致 -> 隐藏
      if (keyValues[k].size > 1) p1.push(k);
    }} else if (cnt >= 2) {{
      p2.push(k);
    }} else {{
      p3.push(k);
    }}
  }});
  // 按出现次数降序 + 名称升序排序，保证同优先级列顺序稳定
  const sortKeys = arr => arr.sort((a, b) => (keyCounts[b] - keyCounts[a]) || a.localeCompare(b, 'zh-Hans-CN'));
  sortKeys(p1); sortKeys(p2); sortKeys(p3);
  return {{ p1, p2, p3, maps }};
}}

function renderCell(value, cls) {{
  const v = value == null ? '' : String(value);
  if (!v.trim()) return `<td class="${{cls}} attr-empty-val">—</td>`;
  return `<td class="${{cls}}">${{escapeHtml(v)}}</td>`;
}}

function renderCards() {{
  const container = document.getElementById('cardsContainer');
  container.innerHTML = '';
  CANDIDATES.forEach((group) => {{
    const items = group.same_items || [];
    const {{ p1, p2, p3, maps }} = classifyAttributes(items);
    const card = document.createElement('div');
    card.className = 'group-card';
    // 摘要徽标
    const summaryParts = [];
    if (p1.length) summaryParts.push(`<span class="sum-p1">核心差异 <b>${{p1.length}}</b></span>`);
    if (p2.length) summaryParts.push(`<span class="sum-p2">部分共有 <b>${{p2.length}}</b></span>`);
    if (p3.length) summaryParts.push(`<span class="sum-p3">独占 <b>${{p3.length}}</b></span>`);
    const diffHint = summaryParts.length
      ? `<span class="diff-hint">${{summaryParts.join(' · ')}}</span>`
      : (items.length > 1 ? '<span class="no-diff">✓ 各同款属性一致（已隐藏共同项）</span>' : '');
    // 主图行
    let html = `<div class="main-image-row">
      <img src="${{group.main_image_url}}" alt="主图${{group.index}}"/>
      <span class="label">图${{group.index}}</span>
      ${{diffHint}}
    </div>`;
    // 表头构造：勾选 / 同款主图 / 标题 + 三个优先级列分组
    const hasAttrCols = (p1.length + p2.length + p3.length) > 0;
    let theadHtml = '';
    if (hasAttrCols) {{
      let groupHeaders = '';
      let keyHeaders = '';
      if (p1.length) {{
        groupHeaders += `<th class="grp-p1" colspan="${{p1.length}}">核心差异（全都有·值不同）</th>`;
        p1.forEach(k => {{ keyHeaders += `<th class="col-p1 col-key">${{escapeHtml(k)}}</th>`; }});
      }}
      if (p2.length) {{
        groupHeaders += `<th class="grp-p2" colspan="${{p2.length}}">部分共有</th>`;
        p2.forEach(k => {{ keyHeaders += `<th class="col-p2 col-key">${{escapeHtml(k)}}</th>`; }});
      }}
      if (p3.length) {{
        groupHeaders += `<th class="grp-p3" colspan="${{p3.length}}">独占</th>`;
        p3.forEach(k => {{ keyHeaders += `<th class="col-p3 col-key">${{escapeHtml(k)}}</th>`; }});
      }}
      theadHtml = `<tr>
        <th rowspan="2">勾选</th>
        <th rowspan="2">同款主图</th>
        <th rowspan="2">标题</th>
        ${{groupHeaders}}
      </tr><tr>${{keyHeaders}}</tr>`;
    }} else {{
      theadHtml = `<tr><th>勾选</th><th>同款主图</th><th>标题</th></tr>`;
    }}
    html += `<div class="table-wrap"><table class="same-items-table">
      <thead>${{theadHtml}}</thead>
      <tbody>`;
    items.forEach((item, iIdx) => {{
      const checked = iIdx === 0 ? 'checked' : '';
      const m = maps[iIdx];
      let cells = '';
      p1.forEach(k => {{ cells += renderCell(m[k], 'cell-p1'); }});
      p2.forEach(k => {{ cells += renderCell(m[k], 'cell-p2'); }});
      p3.forEach(k => {{ cells += renderCell(m[k], 'cell-p3'); }});
      html += `<tr>
        <td class="check-cell"><input type="checkbox" data-group="${{group.index}}" data-seq="${{item.seq}}" ${{checked}} onchange="onCheck(this)"/></td>
        <td class="img-cell"><img src="${{item.image_url}}" alt="同款${{item.seq}}"/></td>
        <td class="title-cell">${{escapeHtml(item.title)}}</td>
        ${{cells}}
      </tr>`;
    }});
    html += `</tbody></table></div>`;
    card.innerHTML = html;
    container.appendChild(card);
  }});
}}

function onCheck(el) {{
  const group = el.dataset.group;
  if (!el.checked) {{
    el.checked = true;
    return;
  }}
  document.querySelectorAll(`input[type="checkbox"][data-group="${{group}}"]`).forEach(cb => {{
    if (cb !== el) cb.checked = false;
  }});
}}

function copySelections() {{
  const lines = [];
  CANDIDATES.forEach((group) => {{
    const checked = document.querySelector(`input[type="checkbox"][data-group="${{group.index}}"]:checked`);
    if (checked) {{
      lines.push(`\u56fe${{group.index}}: \u540c\u6b3e\u5e8f\u53f7${{checked.dataset.seq}}`);
    }}
  }});
  const text = lines.join('\\n');
  copyToClipboard(text);
  showToast('\u2705 \u5df2\u590d\u5236\u52fe\u9009\u7ed3\u679c');
}}

function copyToClipboard(text) {{
  if (navigator.clipboard) {{
    navigator.clipboard.writeText(text);
  }} else {{
    const ta = document.createElement('textarea');
    ta.value = text;
    document.body.appendChild(ta);
    ta.select();
    document.execCommand('copy');
    document.body.removeChild(ta);
  }}
}}

function showToast(msg) {{
  const t = document.getElementById('toast');
  t.textContent = msg;
  t.classList.add('show');
  setTimeout(() => t.classList.remove('show'), 2000);
}}

renderCards();
</script>
</body>
</html>"""

    timestamp = int(time.time() * 1000)
    html_path = os.path.join(HTML_OUTPUT_DIR, f"publish_select_same_{timestamp}.html")
    html_path = os.path.abspath(html_path)
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    # 自动打开 HTML 文件
    try:
        system = platform.system()
        if system == "Darwin":
            subprocess.Popen(["open", html_path])
        elif system == "Windows":
            os.startfile(html_path)
        else:
            subprocess.Popen(["xdg-open", html_path])
    except Exception:
        pass

    return html_path


def main(args=None):
    """
    商品发布命令入口。

    Args:
        args: 命令参数列表。为 None 时从 sys.argv[1:] 读取。
              通过 cli.py 路由时传入 sys.argv[2:]。
    """
    parser = argparse.ArgumentParser(
        prog="product_publish",
        description=COMMAND_DESC,
    )
    parser.add_argument("--image", type=str, default="", help="本地图片路径")
    parser.add_argument("--url", type=str, default="", help="图片链接")
    parser.add_argument("--select", type=int, default=0, help="选择同款商品序号（从1开始），自动读取上次 AI 识图保存的状态")
    parser.add_argument("--publish-data", type=str, default="", help="【可选/向后兼容】上一步 AI 识别返回的数据（JSON字符串），不传则自动读取状态文件")
    # 批量模式参数
    parser.add_argument("--images", nargs="+", default=[], metavar="IMAGE",
                        help="批量本地图片路径，空格分隔多个路径")
    parser.add_argument("--urls", nargs="+", default=[], metavar="URL",
                        help="批量图片链接，空格分隔多个链接")
    parser.add_argument("--batch-select", type=str, default="",
                        help="批量选品序号，逗号分隔，与有同款的图片顺序一致（例如：1,2,1）")
    parser.add_argument("--batch-default-first", action="store_true",
                        help="批量模式下全选每张图片的第一个同款，直接生成发品 Excel")
    parser.add_argument("--show-all-candidates", action="store_true",
                        help="批量模式下展示所有图片的全部同款候选信息，供商家逐张选择")

    if args is None:
        args = sys.argv[1:]
    parsed = parser.parse_args(args)

    try:
        # 冲突检测：批量模式与单图模式不可同时使用
        batch_mode = bool(
            parsed.images or parsed.urls or parsed.batch_select
            or parsed.batch_default_first or parsed.show_all_candidates
        )
        single_mode = bool(parsed.image or parsed.url or parsed.select or parsed.publish_data)
        if batch_mode and single_mode:
            print_output(make_output(
                success=False,
                error_code="PARAM_CONFLICT",
                markdown="❌ 批量模式（--images/--urls/--batch-select/--batch-default-first/--show-all-candidates）与单图模式（--image/--url/--select）不可同时使用。",
            ))
            return

        # 批量模式A：批量上传 + AI 识图
        if parsed.images or parsed.urls:
            pic_urls = []       # 最终图片 URL 列表
            pic_labels = []     # 展示标签（文件名或原始 URL）
            pic_paths = []      # 本地图片的绝对路径（URL 模式下使用 URL 本身），与 pic_urls 顺序一一对应
            upload_errors = []  # 上传失败记录

            # 上传本地图片到图片银行
            for img_path in parsed.images:
                if not os.path.exists(img_path):
                    upload_errors.append(f"`{img_path}`：文件不存在")
                    continue
                try:
                    url = upload_image_to_bank(img_path)
                    pic_urls.append(url)
                    pic_labels.append(os.path.basename(img_path))
                    pic_paths.append(os.path.abspath(img_path))
                except Exception as e:
                    upload_errors.append(f"`{img_path}`：上传失败 — {e}")

            # 直接追加 URL 链接
            for url in parsed.urls:
                pic_urls.append(url)
                pic_labels.append(url)
                pic_paths.append(url)

            if not pic_urls:
                err_detail = "\n".join(f"- {e}" for e in upload_errors)
                print_output(make_output(
                    success=False,
                    error_code="ALL_UPLOAD_FAILED",
                    markdown=f"❌ 所有图片均上传失败，无法继续识图：\n\n{err_detail}",
                ))
                return

            # 对每个 URL 调用 AI 识图
            batch_items = []
            for i, (url, label, pic_path) in enumerate(zip(pic_urls, pic_labels, pic_paths), start=1):
                try:
                    result = ai_publish_by_image(url)
                    batch_items.append({
                        "index": i,
                        "label": label,
                        "picPath": pic_path,
                        "picUrl": url,
                        "categoryId": result.get("categoryId", 0),
                        "categoryName": result.get("categoryName", ""),
                        "tkItemIds": result.get("tkItemIds", []),
                        "effectiveItemIds": result.get("effectiveItemIds", []),
                        "redirect": result.get("redirect"),
                    })
                except Exception as e:
                    batch_items.append({
                        "index": i,
                        "label": label,
                        "picPath": pic_path,
                        "picUrl": url,
                        "categoryId": 0,
                        "categoryName": "",
                        "tkItemIds": [],
                        "effectiveItemIds": [],
                        "redirect": "https://offer-new.1688.com/select.htm",
                        "error": str(e),
                    })

            # 保存批量识图状态
            _save_batch_state(batch_items)

            # 过滤：仅保留有同款商品的图片
            selectable = [
                item for item in batch_items
                if not item.get("redirect") and not item.get("error") and item.get("effectiveItemIds")
            ]
            filtered = [
                item for item in batch_items
                if item.get("redirect") or item.get("error") or not item.get("effectiveItemIds")
            ]

            # 构建展示 Markdown
            total = len(batch_items)
            lines = [f"## 批量识图完成（共 {total} 张，{len(selectable)} 张有同款，{len(filtered)} 张已过滤）\n"]

            if upload_errors:
                lines.append("**⚠️ 以下图片上传失败，已跳过：**")
                for e in upload_errors:
                    lines.append(f"- {e}")
                lines.append("")

            if filtered:
                lines.append("**以下图片未识别到同款商品，已自动过滤（不展示、不生成发品链接）：**")
                for item in filtered:
                    lines.append(f"- 图片 {item['index']}：{item['label']}")
                lines.append("")

            # 仅展示有同款的图片（紧凑模式：每张仅展示首选同款的主图 + 标题 + 属性）
            for seq, item in enumerate(selectable, start=1):
                idx = item["index"]
                label = item["label"]
                effective_items = item.get("effectiveItemIds", [])
                cat_name = item.get("categoryName", "")
                first = effective_items[0]
                first_title = first.get("sameItemTitle", "")
                first_pic = first.get("sameItemPicUrl", "")
                first_attrs = _format_seller_point(first.get("sameItemSellerPoint"))
                lines.append("---\n")
                lines.append(
                    f"### 序号 {seq}｜图片 {idx}：{label}　　类目：{cat_name}　　同款数：{len(effective_items)}\n"
                )
                if first_pic:
                    lines.append(f"![首选同款]({first_pic})\n")
                lines.append(f"**首选同款标题**：{first_title}")
                lines.append(f"**首选同款属性**：{first_attrs}")
                lines.append("")

            # 操作引导：询问“全选第一个 / 查看全部候选”
            if selectable:
                sel_indices = ", ".join(str(item["index"]) for item in selectable)
                lines.append("---\n")
                lines.append("请选择下一步操作：\n")
                lines.append("**1. 默认全选第一个同款（推荐）** — 直接生成发品 Excel：")
                lines.append("```")
                lines.append("product_publish --batch-default-first")
                lines.append("```\n")
                lines.append("**2. 查看所有同款信息后再逐张选择**：")
                lines.append("```")
                lines.append("product_publish --show-all-candidates")
                lines.append("```\n")
                lines.append(f"（有同款的图片共 {len(selectable)} 张，对应图片序号：{sel_indices}）")

            print_output(make_output(
                success=True,
                markdown="\n".join(lines),
                data={
                    "total": total,
                    "items": [
                        {
                            "index": item["index"],
                            "label": item["label"],
                            "candidates": len(item.get("effectiveItemIds", [])),
                            "recognizable": not bool(item.get("redirect") or item.get("error")),
                        }
                        for item in batch_items
                    ],
                },
            ))
            return

        # 批量模式C：展示所有同款候选详情（仅展示，不执行选品）
        if parsed.show_all_candidates:
            state = _load_batch_state()
            if not state:
                print_output(make_output(
                    success=False,
                    error_code="NO_BATCH_STATE",
                    markdown="❌ 未找到批量识图记录，请先执行 `product_publish --images ...` 或 `product_publish --urls ...` 进行批量识图。",
                ))
                return

            items = state.get("items", [])
            selectable_items = [
                item for item in items
                if not item.get("redirect") and not item.get("error") and item.get("effectiveItemIds")
            ]
            filtered_items = [
                item for item in items
                if item.get("redirect") or item.get("error") or not item.get("effectiveItemIds")
            ]

            if not selectable_items:
                print_output(make_output(
                    success=False,
                    error_code="NO_SELECTABLE_ITEMS",
                    markdown="❌ 所有图片均未识别到同款商品，无法展示候选。",
                ))
                return

            lines = [f"## 全部同款候选信息（共 {len(selectable_items)} 张图片有同款）\n"]

            if filtered_items:
                lines.append(f"**已过滤 {len(filtered_items)} 张无同款图片（不要求选择、不生成发品链接）**\n")

            for seq, item in enumerate(selectable_items, start=1):
                idx = item["index"]
                label = item["label"]
                effective_items = item.get("effectiveItemIds", [])
                cat_name = item.get("categoryName", "")
                lines.append("---\n")
                lines.append(f"### 序号 {seq}｜图片 {idx}：{label}　　类目：{cat_name}\n")
                lines.append("| 序号 | 商品标题 | 属性 |")
                lines.append("|------|---------|------|")
                for j, eff in enumerate(effective_items, start=1):
                    title = eff.get("sameItemTitle", "")
                    attrs = _format_seller_point(eff.get("sameItemSellerPoint"))
                    lines.append(f"| {j} | {title} | {attrs} |")
                lines.append("")
                previews = [
                    f"{j}.![商品{j}]({eff.get('sameItemPicUrl', '')})"
                    for j, eff in enumerate(effective_items, start=1)
                    if eff.get("sameItemPicUrl")
                ]
                if previews:
                    lines.append("**主图预览**：" + "　".join(previews))
                lines.append("")

            # 构造 HTML 勾选页所需的 candidates 数据并生成 HTML
            candidates_payload = []
            for seq, item in enumerate(selectable_items, start=1):
                same_items_payload = []
                for j, eff in enumerate(item.get("effectiveItemIds", []), start=1):
                    same_items_payload.append({
                        "seq": j,
                        "image_url": eff.get("sameItemPicUrl", ""),
                        "title": eff.get("sameItemTitle", ""),
                        "attributes": _format_seller_point(eff.get("sameItemSellerPoint")),
                    })
                candidates_payload.append({
                    "index": seq,
                    "main_image_url": item.get("picUrl", ""),
                    "same_items": same_items_payload,
                })

            html_path = _generate_select_same_html(candidates_payload)

            example = ",".join("1" for _ in selectable_items)
            sel_indices = ", ".join(str(item["index"]) for item in selectable_items)
            lines.append("---\n")
            lines.append("✨ **已为您打开《批量发品 - 同款勾选》网页**\n")
            lines.append("- 每张主图默认勾选第 1 个同款，可自由调整")
            lines.append("- 同一张主图分组内只能勾选一个同款（互斥）")
            lines.append("- 完成后点击页面顶部或底部的「📋 复制勾选同款」按钮")
            lines.append("- 复制成功后**返回聊天框粘贴勾选结果**，继续生成发品 Excel\n")
            lines.append(f"> 网页未自动打开？手动打开文件：`{html_path}`\n")
            lines.append("---\n")
            lines.append("如不使用网页勾选，也可直接选择以下命令：\n")
            lines.append("**1. 默认全选第一个同款** — 直接生成发品 Excel：")
            lines.append("```")
            lines.append("product_publish --batch-default-first")
            lines.append("```\n")
            lines.append(f"**2. 自定义每个同款** — 按图片顺序（图片 {sel_indices}）提供逗号分隔的序号：")
            lines.append("```")
            lines.append(f'product_publish --batch-select "{example}"')
            lines.append("```\n")
            lines.append(f"（序号个数必须为 {len(selectable_items)}，每位取值范围 1 ~ 4）")

            print_output(make_output(
                success=True,
                markdown="\n".join(lines),
                data={
                    "selectableCount": len(selectable_items),
                    "htmlPath": html_path,
                    "items": [
                        {
                            "index": item["index"],
                            "label": item["label"],
                            "candidates": len(item.get("effectiveItemIds", [])),
                        }
                        for item in selectable_items
                    ],
                },
            ))
            return

        # 批量模式B：批量选品保存 + 生成发品链接（支持 --batch-select 与 --batch-default-first）
        if parsed.batch_select or parsed.batch_default_first:
            state = _load_batch_state()
            if not state:
                print_output(make_output(
                    success=False,
                    error_code="NO_BATCH_STATE",
                    markdown="❌ 未找到批量识图记录，请先执行 `product_publish --images ...` 或 `product_publish --urls ...` 进行批量识图。",
                ))
                return

            items = state.get("items", [])
            if not items:
                print_output(make_output(
                    success=False,
                    error_code="EMPTY_BATCH_STATE",
                    markdown="❌ 批量识图记录为空，请重新执行批量识图。",
                ))
                return

            # 过滤：仅保留有同款的图片，无同款图片不参与选品、不生成发品链接
            selectable_items = [
                item for item in items
                if not item.get("redirect") and not item.get("error") and item.get("effectiveItemIds")
            ]
            filtered_items = [
                item for item in items
                if item.get("redirect") or item.get("error") or not item.get("effectiveItemIds")
            ]

            if not selectable_items:
                print_output(make_output(
                    success=False,
                    error_code="NO_SELECTABLE_ITEMS",
                    markdown="❌ 所有图片均未识别到同款商品，无法进行选品。",
                ))
                return

            # 解析选品序号：--batch-default-first 等同于全选 1；否则解析 --batch-select
            if parsed.batch_default_first:
                selections = [1] * len(selectable_items)
            else:
                try:
                    selections = [int(s.strip()) for s in parsed.batch_select.split(",")]
                except ValueError:
                    print_output(make_output(
                        success=False,
                        error_code="INVALID_BATCH_SELECT",
                        markdown="❌ 选品序号格式错误，请使用逗号分隔的整数，例如：`1,2,1`",
                    ))
                    return

                # 校验数量：与有同款的图片数量一致
                if len(selections) != len(selectable_items):
                    print_output(make_output(
                        success=False,
                        error_code="SELECTION_COUNT_MISMATCH",
                        markdown=f"❌ 选品数量不匹配：共有 {len(selectable_items)} 张有同款的图片，但提供了 {len(selections)} 个选择序号。",
                    ))
                    return

            # 逐项保存并生成发品链接（仅有同款的图片）
            results = []
            errors = []
            for item, sel in zip(selectable_items, selections):
                idx = item["index"]
                label = item["label"]

                effective_items = item.get("effectiveItemIds", [])

                if sel < 1 or sel > len(effective_items):
                    errors.append(
                        f"图片 {idx}（{label}）：序号 {sel} 超出范围（1 ~ {len(effective_items)}）"
                    )
                    continue

                selected = effective_items[sel - 1]
                try:
                    aigc_time = ai_publish_save(
                        pic_url=item["picUrl"],
                        category_id=item["categoryId"],
                        category_name=item["categoryName"],
                        tk_item_ids=item["tkItemIds"],
                        absolute_same_item_id=selected["sameItemId"],
                        absolute_same_item_title=selected["sameItemTitle"],
                    )
                    publish_url = build_publish_url(aigc_time, item["categoryId"])
                    results.append({
                        "index": idx,
                        "label": label,
                        "picPath": item.get("picPath", label),
                        "picUrl": item.get("picUrl", ""),
                        "publishUrl": publish_url,
                        "selected": selected.get("sameItemTitle", ""),
                        "attributes": _format_seller_point(selected.get("sameItemSellerPoint")),
                    })
                except Exception as e:
                    errors.append(f"图片 {idx}（{label}）：保存选品失败 — {e}")

            if errors:
                err_detail = "\n".join(f"- {e}" for e in errors)
                print_output(make_output(
                    success=False,
                    error_code="BATCH_SAVE_PARTIAL_FAIL",
                    markdown=f"❌ 部分选品保存失败：\n\n{err_detail}\n\n请检查序号后重试。",
                ))
                return

            # 生成 Excel 文件（openpyxl，兼容所有 Excel 版本）
            wb = Workbook()
            ws = wb.active
            ws.title = "批量发品链接"

            # 表头
            headers = ["序号", "本地图片", "同款标题/发品链接", "同款属性"]
            ws.append(headers)
            for col in range(1, 5):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.protection = Protection(locked=False)

            # 填充数据
            for row_idx, r in enumerate(results, start=2):
                # 序号
                num_cell = ws.cell(row=row_idx, column=1, value=row_idx - 1)
                num_cell.alignment = Alignment(horizontal="center", vertical="center")
                num_cell.protection = Protection(locked=False)

                # 列B：TwoCellAnchor 图片，锁定在单元格内
                thumb_path = _prepare_image_for_excel(r["picPath"], r.get("picUrl", ""), row_idx)
                if thumb_path:
                    try:
                        xl_img = XlImage(thumb_path)
                        anchor = TwoCellAnchor()
                        anchor._from = AnchorMarker(col=1, colOff=0, row=row_idx - 1, rowOff=0)
                        anchor.to   = AnchorMarker(col=2, colOff=0, row=row_idx,     rowOff=0)
                        xl_img.anchor = anchor
                        ws.add_image(xl_img)
                        ws.row_dimensions[row_idx].height = _THUMB_ROW_HEIGHT
                    except Exception:
                        pass

                # 列C：超链接标题
                title_cell = ws.cell(row=row_idx, column=3, value=r["selected"])
                title_cell.hyperlink = r["publishUrl"]
                title_cell.font = Font(color="0000FF", underline="single")
                title_cell.alignment = Alignment(vertical="center", wrap_text=True)
                title_cell.protection = Protection(locked=False)

                # 列D：属性
                attr_cell = ws.cell(row=row_idx, column=4, value=r["attributes"])
                attr_cell.alignment = Alignment(vertical="center", wrap_text=True)
                attr_cell.protection = Protection(locked=False)

            # 列宽
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = _IMG_COL_WIDTH
            ws.column_dimensions["C"].width = 50
            ws.column_dimensions["D"].width = 50

            # 工作表保护：锁定图片（不可拖动/缩放），单元格保持可编辑
            ws.protection.sheet = True
            ws.protection.objects = True             # 锁定图片不可操作
            ws.protection.scenarios = False
            ws.protection.selectLockedCells = False   # 锁定单元格仍可点击
            ws.protection.selectUnlockedCells = False # 解锁单元格可正常编辑

            # 保存文件
            now = datetime.now()
            expire = now + timedelta(days=3)
            now_str = f"{now.year}年{now.month}月{now.day}日{now.strftime('%H:%M')}"
            expire_str = f"{expire.year}年{expire.month}月{expire.day}日{expire.strftime('%H:%M')}"
            excel_filename = f"批量发品链接_{now_str}_过期{expire_str}.xlsx"
            excel_path = os.path.join(os.path.dirname(__file__), "..", "..", excel_filename)
            excel_path = os.path.abspath(excel_path)
            wb.save(excel_path)

            # 用系统默认程序打开 Excel 文件
            try:
                system = platform.system()
                if system == "Darwin":
                    subprocess.Popen(["open", excel_path])
                elif system == "Windows":
                    os.startfile(excel_path)
                else:
                    subprocess.Popen(["xdg-open", excel_path])
            except Exception:
                pass

            # 构建输出 Markdown
            lines = [f"## 批量发品完成（共 {len(results)} 条）\n"]

            if filtered_items:
                lines.append(f"**已过滤 {len(filtered_items)} 张无同款图片（不生成发品链接）**\n")

            lines.append(f"发品链接已生成到 Excel 文件：\n")
            lines.append(f"```\n{excel_path}\n```\n")
            lines.append("请打开 Excel 文件查看发品链接，复制链接到浏览器中打开，补充商品信息（价格、库存、SKU 等）并点击【发布商品】按钮。")
            lines.append(
                "\n⏰ **发品链接有效期为 3 天，请及时完成发布。**"
            )
            lines.append(
                "\n发布成功后还可以使用 **AI 标题优化**、**AI 主图优化** 进一步提升商品质量。"
            )

            print_output(make_output(
                success=True,
                markdown="\n".join(lines),
                data={
                    "excelPath": excel_path,
                    "publishLinks": [
                        {
                            "index": r["index"],
                            "label": r["label"],
                            "picPath": r["picPath"],
                            "publishUrl": r["publishUrl"],
                        }
                        for r in results
                    ]
                },
            ))
            return

        # 模式1：本地图片上传
        if parsed.image:
            if not os.path.exists(parsed.image):
                print_output(make_output(
                    success=False,
                    error_code="FILE_NOT_FOUND",
                    markdown=f"❌ 图片文件不存在：`{parsed.image}`",
                ))
                return

            pic_url = upload_image_to_bank(parsed.image)
            print_output(make_output(
                success=True,
                markdown=f"## 图片上传成功\n\n图片已上传至图片银行。\n\n**图片URL**：\n\n![上传图片]({pic_url})",
                data={"picUrl": pic_url},
            ))
            return

        # 模式2：AI 识别同款
        if parsed.url and not parsed.select and not parsed.publish_data:
            result = ai_publish_by_image(parsed.url)

            if "redirect" in result:
                print_output(make_output(
                    success=True,
                    markdown=f"## AI 识图结果\n\n未识别到同款商品，请前往 [1688 发品页面]({result['redirect']}) 手动发布。",
                ))
                return

            effective_items = result.get("effectiveItemIds", [])
            if not effective_items:
                print_output(make_output(
                    success=True,
                    markdown="## AI 识图结果\n\n未找到同款商品，请尝试其他图片。",
                ))
                return

            # 自动保存状态，供后续 --select 直接使用
            _save_state({
                "picUrl": parsed.url,
                "categoryId": result.get("categoryId", 0),
                "categoryName": result.get("categoryName", ""),
                "tkItemIds": result.get("tkItemIds", []),
                "effectiveItemIds": result.get("effectiveItemIds", []),
            })

            lines = ["## AI 识别到以下同款商品\n"]
            for idx, item in enumerate(effective_items, start=1):
                title = item.get("sameItemTitle", "")
                pic_url_item = item.get("sameItemPicUrl", "")
                seller_point = _format_seller_point(item.get("sameItemSellerPoint"))
                lines.append(f"**商品{idx}**")
                lines.append(f"- 标题：{title}")
                if pic_url_item:
                    lines.append(f"- 主图：![商品{idx}]({pic_url_item})")
                lines.append(f"- 属性：{seller_point}")
                lines.append("")

            lines.append("请选择一个同款商品（输入商品编号）")

            # data 字段脱敏：仅保留展示所需信息，不暴露内部 ID
            safe_data = {
                "items": [
                    {
                        "title": item.get("sameItemTitle", ""),
                        "picUrl": item.get("sameItemPicUrl", ""),
                        "attributes": item.get("sameItemSellerPoint", {}),
                    }
                    for item in effective_items
                ]
            }
            print_output(make_output(
                success=True,
                markdown="\n".join(lines),
                data=safe_data,
            ))
            return

        # 模式3：用户选品保存
        if parsed.select:
            # 优先使用 --publish-data（向后兼容），否则自动读取状态文件
            if parsed.publish_data:
                try:
                    publish_data = json.loads(parsed.publish_data)
                except json.JSONDecodeError:
                    print_output(make_output(
                        success=False,
                        error_code="INVALID_JSON",
                        markdown="❌ 传入的数据格式不正确，请检查后重试。",
                    ))
                    return
                pic_url = parsed.url
            else:
                state = _load_state()
                if not state:
                    print_output(make_output(
                        success=False,
                        error_code="NO_SAVED_STATE",
                        markdown="❌ 未找到 AI 识图记录，请先上传图片或提供图片链接进行识图。",
                    ))
                    return
                publish_data = state
                pic_url = state.get("picUrl", "")

            effective_items = publish_data.get("effectiveItemIds", [])
            if not effective_items:
                print_output(make_output(
                    success=False,
                    error_code="NO_EFFECTIVE_ITEMS",
                    markdown="❌ 未找到可选择的同款商品，请重新进行 AI 识图。",
                ))
                return

            select_idx = parsed.select
            if select_idx < 1 or select_idx > len(effective_items):
                print_output(make_output(
                    success=False,
                    error_code="INVALID_SELECT",
                    markdown=f"❌ 商品序号 `{select_idx}` 无效，请输入 1 ~ {len(effective_items)} 之间的数字",
                ))
                return

            selected_item = effective_items[select_idx - 1]
            absolute_same_item_id = selected_item.get("sameItemId", "")
            absolute_same_item_title = selected_item.get("sameItemTitle", "")

            if not absolute_same_item_id:
                print_output(make_output(
                    success=False,
                    error_code="MISSING_ITEM_ID",
                    markdown="❌ 所选商品信息不完整，请重新选择。",
                ))
                return

            category_id = publish_data.get("categoryId", 0)
            category_name = publish_data.get("categoryName", "")
            tk_item_ids = publish_data.get("tkItemIds", [])

            aigc_time = ai_publish_save(
                pic_url=pic_url,
                category_id=category_id,
                category_name=category_name,
                tk_item_ids=tk_item_ids,
                absolute_same_item_id=absolute_same_item_id,
                absolute_same_item_title=absolute_same_item_title,
            )

            publish_url = build_publish_url(aigc_time, category_id)

            print_output(make_output(
                success=True,
                markdown=(
                    f"## 选品保存成功\n\n"
                    f"已保存选品数据，请复制以下链接到浏览器中打开，进入 1688 发品页面补充商品信息（价格、库存、SKU 等）并点击【发布商品】按钮：\n\n"
                    f"[👉 前往 1688 发品页面]({publish_url})\n\n"
                    f"**发品链接**：\n\n"
                    f"```\n{publish_url}\n```"
                ),
                data={
                    "aigcTime": aigc_time,
                    "catId": category_id,
                    "publishUrl": publish_url,
                },
            ))
            return

        # 无匹配模式：输出引导
        print_output(make_output(
            success=False,
            error_code="MISSING_PARAM",
            markdown="""## 参数错误

请按以下方式使用：

### 单图模式

1. **上传本地图片**：
   ```
   python3 cli.py product_publish --image /path/to/image.jpg
   ```

2. **AI 识图（图片链接）**：
   ```
   python3 cli.py product_publish --url https://example.com/image.jpg
   ```

3. **选择同款并保存**（状态自动保持，无需手动传 JSON）：
   ```
   python3 cli.py product_publish --select 1
   ```
   如需手动传入数据（向后兼容）：
   ```
   python3 cli.py product_publish --url <picUrl> --select 1 --publish-data '<JSON>'
   ```

### 批量模式

4. **批量上传本地图片 + AI 识图**：
   ```
   python3 cli.py product_publish --images img1.jpg img2.jpg img3.jpg
   ```

5. **批量图片链接 AI 识图**：
   ```
   python3 cli.py product_publish --urls url1 url2 url3
   ```

6. **混合模式（本地图片 + 链接）**：
   ```
   python3 cli.py product_publish --images img1.jpg --urls url2 url3
   ```

7. **查看全部同款候选详情**（仅展示，不执行选品）：
   ```
   python3 cli.py product_publish --show-all-candidates
   ```

8. **默认全选第一个同款并生成发品 Excel**：
   ```
   python3 cli.py product_publish --batch-default-first
   ```

9. **自定义选品并生成所有发品链接**：
   ```
   python3 cli.py product_publish --batch-select "1,2,1"
   ```
   （逗号分隔，每位对应一张有同款的图片的选择序号）
""",
        ))

    except Exception as e:
        print_error(e)


if __name__ == "__main__":
    main()
